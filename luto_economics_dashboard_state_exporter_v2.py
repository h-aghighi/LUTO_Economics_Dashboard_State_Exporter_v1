#!/usr/bin/env python3
"""
luto_economics_dashboard_state_exporter_v2.py

Purpose
-------
Extract LUTO dashboard economics overview tables and export ONLY final Excel workbooks.

This version supports two modes:

1. Single-run mode
   Use --data-dir pointing to one run's DATA_REPORT/data folder.

2. Batch-run mode
   Use --reports-base-dir pointing to a folder that contains multiple Run_* folders.

Example batch input:

    S:\...\Report_Data
      Run_G0001\DATA_REPORT\data\Economics_overview_sum.js
      Run_G0002\DATA_REPORT\data\Economics_overview_sum.js
      Run_G0013\DATA_REPORT\data\Economics_overview_sum.js

Batch output:

    outputs/
      Queensland/
        Run_G0001_Queensland_Economics_Dashboard_Final_Table.xlsx
        Run_G0002_Queensland_Economics_Dashboard_Final_Table.xlsx
        Run_G0013_Queensland_Economics_Dashboard_Final_Table.xlsx
      Victoria/
        Run_G0001_Victoria_Economics_Dashboard_Final_Table.xlsx
        ...

Each workbook contains one sheet per region in that state.

Requirements
------------
python -m pip install pandas openpyxl
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: Missing packages. Run: python -m pip install pandas openpyxl")
    raise


DASHBOARD_FILE = "Economics_overview_sum.js"

DASHBOARD_SERIES_ORDER = [
    "Agricultural Land-use (revenue)",
    "Agricultural Management (revenue)",
    "Non-Agricultural Land-use (revenue)",
    "Agricultural Land-use (cost)",
    "Agricultural Management (cost)",
    "Non-Agricultural Land-use (cost)",
    "Transition cost (Ag2Ag)",
    "Transition cost (Ag2Non-Ag)",
    "Profit",
]

# Default NRM/state mapping for regions visible in current LUTO dashboard exports.
# Unknown names are written to Unknown_State and printed at the end.
REGION_TO_STATE = {
    # National / territories
    "AUSTRALIA": "National",
    "ACT": "Australian Capital Territory",
    "Northern Territory": "Northern Territory",

    # Queensland
    "Burdekin": "Queensland",
    "Burnett Mary": "Queensland",
    "Cape York": "Queensland",
    "Condamine": "Queensland",
    "Desert Channels": "Queensland",
    "Fitzroy": "Queensland",
    "Mackay Whitsunday": "Queensland",
    "Maranoa Balonne and Border Rivers": "Queensland",
    "Northern Gulf": "Queensland",
    "South East Queensland": "Queensland",
    "South West Queensland": "Queensland",
    "Southern Gulf": "Queensland",
    "Torres Strait": "Queensland",
    "Wet Tropics": "Queensland",

    # New South Wales
    "Central Tablelands": "New South Wales",
    "Central West": "New South Wales",
    "Greater Sydney": "New South Wales",
    "Hunter": "New South Wales",
    "Murray": "New South Wales",
    "North Coast": "New South Wales",
    "North West NSW": "New South Wales",
    "Northern Tablelands": "New South Wales",
    "Riverina": "New South Wales",
    "South East NSW": "New South Wales",
    "Western": "New South Wales",

    # Victoria
    "Corangamite": "Victoria",
    "East Gippsland": "Victoria",
    "Glenelg Hopkins": "Victoria",
    "Goulburn Broken": "Victoria",
    "Mallee": "Victoria",
    "North Central": "Victoria",
    "North East": "Victoria",
    "Port Phillip and Western Port": "Victoria",
    "West Gippsland": "Victoria",
    "Wimmera": "Victoria",

    # South Australia
    "Adelaide and Mount Lofty Ranges": "South Australia",
    "Alinytjara Wilurara": "South Australia",
    "Eyre Peninsula": "South Australia",
    "Kangaroo Island": "South Australia",
    "Northern and Yorke": "South Australia",
    "South Australian Arid Lands": "South Australia",
    "South Australian Murray Darling Basin": "South Australia",
    "South East": "South Australia",

    # Western Australia
    "Avon River Basin": "Western Australia",
    "Northern Agricultural Region": "Western Australia",
    "Peel-Harvey Region": "Western Australia",
    "Rangelands Region": "Western Australia",
    "South Coast Region": "Western Australia",
    "South West Region": "Western Australia",
    "Swan Region": "Western Australia",

    # Tasmania
    "North NRM Region": "Tasmania",
    "North West NRM Region": "Tasmania",
    "South NRM Region": "Tasmania",
}


def log(msg: str) -> None:
    print(msg, flush=True)


def sanitize_filename(s: Any) -> str:
    out = re.sub(r'[*?:"<>|\\/]+', "_", str(s))
    out = re.sub(r"\s+", "_", out).strip("_ .")
    return out or "output"


def sanitize_sheet_name(s: Any) -> str:
    out = re.sub(r'[\[\]\*\?:/\\]', "_", str(s))
    out = re.sub(r"\s+", " ", out).strip(" .")
    return (out or "Sheet")[:31]


def unique_sheet_name(name: Any, used: set[str]) -> str:
    base = sanitize_sheet_name(name)
    candidate = base
    i = 1
    while candidate.lower() in used:
        suffix = f"_{i}"
        candidate = (base[: 31 - len(suffix)] + suffix)[:31]
        i += 1
    used.add(candidate.lower())
    return candidate


def fix_data_dir(data_dir: Path) -> Path:
    if data_dir.name.lower() == "map_layers" and data_dir.parent.name.lower() == "data":
        log("WARNING: You supplied DATA_REPORT/data/map_layers.")
        log("         Switching to the parent DATA_REPORT/data folder:")
        log(f"         {data_dir.parent}")
        return data_dir.parent
    return data_dir


def read_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="ignore")


def strip_comments_and_trailing_commas(s: str) -> str:
    out: List[str] = []
    i = 0
    in_str: Optional[str] = None
    escape = False
    while i < len(s):
        ch = s[i]
        nxt = s[i + 1] if i + 1 < len(s) else ""
        if in_str:
            out.append(ch)
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == in_str:
                in_str = None
            i += 1
            continue
        if ch in ('"', "'"):
            in_str = ch
            out.append(ch)
            i += 1
            continue
        if ch == "/" and nxt == "/":
            i += 2
            while i < len(s) and s[i] not in "\r\n":
                i += 1
            continue
        if ch == "/" and nxt == "*":
            i += 2
            while i + 1 < len(s) and not (s[i] == "*" and s[i + 1] == "/"):
                i += 1
            i += 2
            continue
        out.append(ch)
        i += 1
    out_s = "".join(out)
    return re.sub(r",\s*([}\]])", r"\1", out_s)


def js_like_to_json(s: str) -> str:
    s = strip_comments_and_trailing_commas(s)

    def repl_single(m: re.Match) -> str:
        val = m.group(1).replace('"', '\\"')
        return f'"{val}"'

    s = re.sub(r"'([^'\\]*(?:\\.[^'\\]*)*)'", repl_single, s)
    s = re.sub(r"([\{,])\s*([A-Za-z_$][\w$\-+]*)\s*:", r'\1 "\2":', s)
    s = re.sub(r"\bNaN\b", "null", s)
    s = re.sub(r"\bInfinity\b", "null", s)
    s = re.sub(r"\bundefined\b", "null", s)
    return s


def find_window_assignment(js_text: str) -> Tuple[Optional[str], Optional[str]]:
    pattern = re.compile(
        r"window(?:\[['\"](?P<qname>[^'\"]+)['\"]\]|\.(?P<dname>[A-Za-z_$][\w$]*))\s*=\s*",
        re.MULTILINE,
    )
    m = pattern.search(js_text)
    if not m:
        return None, None
    varname = m.group("qname") or m.group("dname")
    start = m.end()
    obj_start = None
    opener = None
    for i in range(start, len(js_text)):
        if js_text[i] in "[{":
            obj_start = i
            opener = js_text[i]
            break
        if js_text[i] == ";":
            break
    if obj_start is None or opener is None:
        return varname, None
    closer = "}" if opener == "{" else "]"
    depth = 0
    in_str: Optional[str] = None
    escape = False
    i = obj_start
    while i < len(js_text):
        ch = js_text[i]
        if in_str:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == in_str:
                in_str = None
            i += 1
            continue
        if ch in ('"', "'"):
            in_str = ch
            i += 1
            continue
        if ch == opener:
            depth += 1
        elif ch == closer:
            depth -= 1
            if depth == 0:
                return varname, js_text[obj_start : i + 1]
        i += 1
    return varname, None


def parse_js(path: Path) -> Tuple[Optional[str], Optional[Any], Optional[str]]:
    txt = read_text(path)
    varname, obj_text = find_window_assignment(txt)
    if obj_text is None:
        return varname, None, "No window[...] assignment found"
    try:
        return varname, json.loads(obj_text), None
    except Exception as e1:
        try:
            return varname, json.loads(js_like_to_json(obj_text)), None
        except Exception as e2:
            return varname, None, f"Parse failed: {type(e1).__name__}; fallback: {type(e2).__name__}"


def looks_like_year(x: Any) -> bool:
    try:
        y = int(float(x))
        return 1900 <= y <= 2200
    except Exception:
        return False


def to_year(x: Any) -> Optional[int]:
    try:
        return int(float(x))
    except Exception:
        return None


def to_float(x: Any) -> Optional[float]:
    try:
        if x is None or x == "":
            return None
        return float(x)
    except Exception:
        return None


def is_series(obj: Any) -> bool:
    if not isinstance(obj, dict) or "data" not in obj:
        return False
    data = obj.get("data")
    if not isinstance(data, list):
        return False
    if not data:
        return True
    return isinstance(data[0], list) and len(data[0]) >= 2 and looks_like_year(data[0][0])


def walk_region_series(obj: Any, rows: List[Dict[str, Any]], region: str) -> None:
    """Recursively extract chart series rows for one region."""
    if is_series(obj):
        series_name = str(obj.get("name", "value"))
        for pair in obj.get("data", []) or []:
            if not isinstance(pair, list) or len(pair) < 2:
                continue
            y = to_year(pair[0])
            v = to_float(pair[1])
            if y is None:
                continue
            rows.append({"region": region, "year": y, "series_name": series_name, "value": v})
        return
    if isinstance(obj, list):
        for item in obj:
            walk_region_series(item, rows, region)
    elif isinstance(obj, dict):
        for v in obj.values():
            walk_region_series(v, rows, region)


def make_region_chart_table(region_rows: pd.DataFrame, start_year: Optional[int], end_year: Optional[int]) -> pd.DataFrame:
    if region_rows.empty:
        return pd.DataFrame(columns=["Category"] + DASHBOARD_SERIES_ORDER)
    work = region_rows.copy()
    work["year"] = pd.to_numeric(work["year"], errors="coerce")
    if start_year is not None:
        work = work[work["year"] >= start_year]
    if end_year is not None:
        work = work[work["year"] <= end_year]
    if work.empty:
        return pd.DataFrame(columns=["Category"] + DASHBOARD_SERIES_ORDER)
    table = work.pivot_table(index="year", columns="series_name", values="value", aggfunc="sum").reset_index()
    table.columns = [str(c) for c in table.columns]
    table = table.rename(columns={"year": "Category"})
    table["Category"] = pd.to_numeric(table["Category"], errors="coerce").astype("Int64")

    observed = [c for c in table.columns if c != "Category"]
    ordered = [s for s in DASHBOARD_SERIES_ORDER if s in observed] + [s for s in observed if s not in DASHBOARD_SERIES_ORDER]
    table = table[["Category"] + ordered].sort_values("Category")
    table = table.astype(object).where(pd.notna(table), "")
    return table


def write_chart_table_sheet(ws, chart_df: pd.DataFrame, chart_title: str) -> None:
    headers = list(chart_df.columns)
    ncols = max(len(headers), 1)

    ws.cell(row=1, column=1, value=chart_title)
    if ncols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=1, column=1).font = Font(bold=False)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left")

    for r_idx, row in enumerate(chart_df.itertuples(index=False, name=None), start=3):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=None if val == "" else val)
            if c_idx == 1:
                cell.font = Font(bold=True)
            if isinstance(cell.value, float):
                cell.number_format = "0.###############"

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ncols):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A3"

    for col_idx in range(1, ncols + 1):
        max_len = 0
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):
            for c in cell:
                max_len = max(max_len, len(str(c.value)) if c.value is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 9), 34)


def load_state_mapping(csv_path: Optional[Path]) -> Dict[str, str]:
    mapping = dict(REGION_TO_STATE)
    if csv_path is None:
        return mapping
    if not csv_path.exists():
        raise FileNotFoundError(f"State mapping CSV not found: {csv_path}")
    custom = pd.read_csv(csv_path)
    required = {"region", "state"}
    if not required.issubset({c.lower() for c in custom.columns}):
        raise ValueError("State mapping CSV must contain columns: region,state")
    cols = {c.lower(): c for c in custom.columns}
    for _, row in custom.iterrows():
        region = str(row[cols["region"]]).strip()
        state = str(row[cols["state"]]).strip()
        if region and state:
            mapping[region] = state
    return mapping


def discover_run_data_dirs(reports_base_dir: Path, run_prefix: str, run_names: List[str]) -> List[Tuple[str, Path]]:
    """Find run folders and their DATA_REPORT/data directories."""
    if not reports_base_dir.exists():
        raise FileNotFoundError(f"Reports base folder does not exist: {reports_base_dir}")

    wanted = {r.lower() for r in run_names}
    found: List[Tuple[str, Path]] = []

    for child in sorted(reports_base_dir.iterdir()):
        if not child.is_dir():
            continue
        if run_prefix and not child.name.startswith(run_prefix):
            continue
        if wanted and child.name.lower() not in wanted:
            continue

        data_dir = child / "DATA_REPORT" / "data"
        if data_dir.exists():
            found.append((child.name, data_dir))
        else:
            log(f"WARNING: Skipping {child.name}; missing DATA_REPORT/data")

    return found


def build_run_list(args: argparse.Namespace) -> List[Tuple[str, Path]]:
    """Build a list of (run_name, data_dir) pairs from single-run or batch inputs."""
    if args.reports_base_dir:
        return discover_run_data_dirs(
            Path(args.reports_base_dir),
            run_prefix=args.run_prefix,
            run_names=args.run_names or [],
        )

    if args.data_dir:
        data_dir = fix_data_dir(Path(args.data_dir))
        # Infer run name from .../Run_G0001/DATA_REPORT/data when possible.
        run_name = args.output_prefix
        parts = list(data_dir.parts)
        if len(parts) >= 3 and data_dir.name.lower() == "data":
            try:
                idx = [p.lower() for p in parts].index("data_report")
                if idx > 0:
                    run_name = parts[idx - 1]
            except ValueError:
                pass
        return [(sanitize_filename(run_name), data_dir)]

    raise ValueError("You must provide either --data-dir for one run or --reports-base-dir for multiple runs.")


def process_one_run(
    run_name: str,
    data_dir: Path,
    output_dir: Path,
    mapping: Dict[str, str],
    wanted_regions: set[str],
    wanted_states: set[str],
    start_year: Optional[int],
    end_year: Optional[int],
    chart_title_template: str,
    include_national: bool,
) -> int:
    """Export all state workbooks for one run."""
    data_dir = fix_data_dir(data_dir)
    prefix = sanitize_filename(run_name)

    if not data_dir.exists():
        log(f"ERROR: [{run_name}] data folder does not exist: {data_dir}")
        return 2

    econ_file = data_dir / DASHBOARD_FILE
    if not econ_file.exists():
        matches = list(data_dir.rglob(DASHBOARD_FILE))
        if not matches:
            log(f"ERROR: [{run_name}] Could not find {DASHBOARD_FILE} under: {data_dir}")
            return 3
        econ_file = matches[0]

    var, parsed, err = parse_js(econ_file)
    if parsed is None or not isinstance(parsed, dict):
        log(f"ERROR: [{run_name}] Could not parse {econ_file}")
        if err:
            log(err)
        return 4

    by_state: Dict[str, List[Tuple[str, pd.DataFrame]]] = {}
    unknown_regions: List[str] = []

    for region in sorted(parsed.keys()):
        if region == "AUSTRALIA" and not include_national:
            continue
        if wanted_regions and region.lower() not in wanted_regions:
            continue

        state = mapping.get(region, "Unknown_State")
        if state == "Unknown_State":
            unknown_regions.append(region)
        if wanted_states and state.lower() not in wanted_states:
            continue

        rows: List[Dict[str, Any]] = []
        walk_region_series(parsed[region], rows, region)
        region_df = pd.DataFrame(rows)
        chart_df = make_region_chart_table(region_df, start_year, end_year)
        if chart_df.empty:
            continue
        by_state.setdefault(state, []).append((region, chart_df))

    if not by_state:
        log(f"ERROR: [{run_name}] No region tables were created after filters.")
        return 5

    for state, region_tables in sorted(by_state.items()):
        state_folder = output_dir / sanitize_filename(state)
        state_folder.mkdir(parents=True, exist_ok=True)
        workbook_path = state_folder / f"{prefix}_{sanitize_filename(state)}_Economics_Dashboard_Final_Table.xlsx"

        wb = Workbook()
        wb.remove(wb.active)
        used_sheet_names: set[str] = set()

        for region, chart_df in region_tables:
            sheet_name = unique_sheet_name(region, used_sheet_names)
            ws = wb.create_sheet(sheet_name)
            title = chart_title_template.format(region=region, state=state, run=run_name)
            write_chart_table_sheet(ws, chart_df, chart_title=title)

        wb.save(workbook_path)
        log(f"Saved: {workbook_path} ({len(region_tables)} region sheet(s))")

    if unknown_regions:
        log(f"WARNING: [{run_name}] These regions were not in the built-in state mapping and were written to Unknown_State:")
        for r in unknown_regions:
            log(f"  - {r}")
        log("You can supply --state-map-csv with columns region,state to override this.")

    return 0


def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(
        description="Write final LUTO economics dashboard tables grouped by state folder and region sheet."
    )

    input_group = parser.add_mutually_exclusive_group(required=True)
    input_group.add_argument(
        "--data-dir",
        help="Single-run mode: path to one run's DATA_REPORT/data folder. If map_layers is supplied, parent data folder is used automatically.",
    )
    input_group.add_argument(
        "--reports-base-dir",
        help="Batch mode: path to folder containing Run_* folders, e.g. .../Report_Data",
    )

    parser.add_argument("--output-dir", required=True, help="Output folder.")
    parser.add_argument(
        "--output-prefix",
        default="",
        help="Single-run mode only: output prefix. If omitted, inferred from the Run_* folder name.",
    )
    parser.add_argument("--run-prefix", default="Run_", help="Batch mode: only process folders starting with this prefix. Default: Run_")
    parser.add_argument("--run-names", nargs="*", default=[], help="Batch mode: optional list of specific run folder names to process.")
    parser.add_argument("--regions", nargs="*", default=[], help="Optional region filter. If omitted, all regions are exported.")
    parser.add_argument("--states", nargs="*", default=[], help="Optional state filter, e.g. Queensland Victoria")
    parser.add_argument("--start-year", type=int, default=2020)
    parser.add_argument("--end-year", type=int, default=2050)
    parser.add_argument(
        "--chart-title-template",
        default="Economics overview for {region}",
        help="Sheet title. Available fields: {region}, {state}, {run}",
    )
    parser.add_argument("--state-map-csv", default="", help="Optional CSV with columns region,state to override/add state mapping.")
    parser.add_argument("--include-national", action="store_true", help="Also export AUSTRALIA into a National folder/workbook.")
    args = parser.parse_args(argv)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    mapping = load_state_mapping(Path(args.state_map_csv) if args.state_map_csv else None)
    wanted_regions = {r.lower() for r in args.regions}
    wanted_states = {s.lower() for s in args.states}

    try:
        run_items = build_run_list(args)
    except Exception as exc:
        log(f"ERROR: {exc}")
        return 1

    if not run_items:
        log("ERROR: No run folders/data folders were found.")
        return 1

    log(f"Found {len(run_items)} run(s) to process.")

    failures = 0
    for run_name, data_dir in run_items:
        log("")
        log("=" * 80)
        log(f"Processing run: {run_name}")
        log(f"Data folder: {data_dir}")
        log("=" * 80)
        code = process_one_run(
            run_name=run_name,
            data_dir=data_dir,
            output_dir=output_dir,
            mapping=mapping,
            wanted_regions=wanted_regions,
            wanted_states=wanted_states,
            start_year=args.start_year,
            end_year=args.end_year,
            chart_title_template=args.chart_title_template,
            include_national=args.include_national,
        )
        if code != 0:
            failures += 1

    if failures:
        log("")
        log(f"Finished with {failures} failed run(s). Successful runs were still exported.")
        return 10

    log("")
    log("Finished. Only final state workbooks were written.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
